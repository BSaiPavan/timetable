import os
import logging
import json
import csv
from flask import Flask, render_template, request, redirect, url_for, jsonify
import io
from flask import send_file
# Custom modules
from config import CONFIG
from solver import generate_timetable, generate_timetable_with_retry
from adapter import build_solver_inputs_from_classes
from extractor import get_solver_data_from_pdf 

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.abspath('uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ── Check for OR-Tools on startup ────────────────────────────────────────────
try:
    from ortools.sat.python import cp_model as _cp_test
    print("✅ OR-Tools available — using CP-SAT solver (fast)")
except ImportError:
    print("⚠️  OR-Tools NOT installed. Falling back to slow backtracking solver.")
    print("   To fix: run  pip install ortools  and restart the app.")
    print("   Without OR-Tools, solving may take 3-5 minutes or fail on large inputs.")

# ── Clear stale session files on every app startup ───────────────────────────
for _f in ["generated_timetable.json", "generated_metadata.json",
           "final_schedule.json", "temp_web_data.json",
           "last_extraction.json", "solver_input_debug.json"]:
    try:
        if os.path.exists(_f):
            os.remove(_f)
    except Exception:
        pass


@app.route("/")
def home():
    return render_template("upload.html")

@app.route("/upload-pdf", methods=["POST"])
def upload_pdf():
    file = request.files.get('file')
    if not file or file.filename == '':
        return "No file selected", 400

    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], "uploaded_schedule.pdf")
    file.save(pdf_path)
    
    try:
        raw_data = get_solver_data_from_pdf(pdf_path) 
        with open("last_extraction.json", "w") as f:
            json.dump(raw_data, f)
        
        CONFIG["raw_extraction"] = raw_data
        return redirect(url_for('generate'))
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return f"AI Extraction Failed: {str(e)}", 500

@app.route("/generate")
def generate():
    data = CONFIG.get("raw_extraction")
    if not data and os.path.exists("last_extraction.json"):
        with open("last_extraction.json", "r") as f:
            data = json.load(f)
    
    if not data:
        return "<h3>No data found. Please upload a PDF first.</h3>"

    try:
        display_data = []
        teacher_map = data.get('teacher_list', {})

        # 1. Process Theory (Now a LIST, not a DICT)
        for class_id, teachers_list in data.get('class_teacher_periods', {}).items():
            # FIX: Loop through the list directly instead of using .items()
            for item in teachers_list:
                t_id = str(item.get('teacher_id'))
                subj = item.get('subject', 'Theory')
                p_val = item.get('periods', 0)

                display_data.append({
                    "class": f"Class {class_id}",
                    "subject": subj,
                    "teacher": teacher_map.get(t_id, {}).get('Name', f"S{t_id}"),
                    "type": "Theory",
                    "periods": p_val,
                    "split_children_json": "[]"
                })

        # 2. Process Labs (Now a LIST, not a DICT)
        for class_id, labs_list in data.get('lab_teacher_periods', {}).items():
            # FIX: Loop through the list directly instead of using .items()
            for item in labs_list:
                t_id = str(item.get('teacher_id'))
                subj = item.get('subject', 'Lab')
                p_raw = item.get('periods', [0])
                p_count = p_raw[0] if isinstance(p_raw, list) else p_raw

                display_data.append({
                    "class": f"Class {class_id}",
                    "subject": subj,
                    "teacher": teacher_map.get(t_id, {}).get('Name', f"S{t_id}"),
                    "type": "Lab",
                    "periods": p_count,
                    "split_children_json": "[]"
                })

        # Extract days/periods from the PDF data if the extractor provided them
        extracted_days    = int(data.get('days', 6))
        extracted_periods = int(data.get('periods', 6))

        return render_template("view_simple.html", rows=display_data,
                               extracted_days=extracted_days,
                               extracted_periods=extracted_periods,
                               merge_groups_json="[]")

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return f"<h3>Data Processing Error: {str(e)}</h3>"
















# ─────────────────────────────────────────────────────────────────────────────
#  HELPER — resolve one cell value from the timetable
#  Timetable structure: timetable[slot_index][class_idx]
#  where slot_index = day * periods_per_day + period
# ─────────────────────────────────────────────────────────────────────────────
def _cell_text(timetable, class_idx, day, period, periods_per_day):
    slot_index = day * periods_per_day + period
    try:
        slot_row = timetable[slot_index]
        # slot_row is either a list [cls0_val, cls1_val, ...] or a dict
        if isinstance(slot_row, list):
            raw = slot_row[class_idx]
        elif isinstance(slot_row, dict):
            raw = slot_row.get(str(class_idx), slot_row.get(class_idx, ""))
        else:
            raw = slot_row
    except (IndexError, KeyError, TypeError):
        return "", "normal"

    if raw is None or raw == 0 or raw == "0":
        return "", "normal"

    text = str(raw).strip()
    lower = text.lower()

    if lower == "free" or lower == "0" or text == "0":
        return "Free", "free"
    if "lab" in lower:
        return text, "lab"
    return text, "normal"


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/download/excel")
def download_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    if not os.path.exists("generated_timetable.json") or \
       not os.path.exists("generated_metadata.json") or \
       not os.path.exists("temp_web_data.json"):
        return "No timetable found. Please generate one first.", 404

    with open("generated_timetable.json")  as f: timetable   = json.load(f)
    with open("generated_metadata.json")   as f: meta        = json.load(f)
    with open("temp_web_data.json")        as f: stored      = json.load(f)

    days        = meta["days"]
    periods     = meta["periods"]
    num_classes = meta["num_classes"]

    organized_keys = list(stored.get("organized", {}).keys())
    all_class_names = [organized_keys[i] if i < len(organized_keys) else str(i+1) for i in range(num_classes)]

    # Support single-class export via ?class_idx=N
    single_idx = request.args.get("class_idx", None)
    if single_idx is not None:
        try:
            single_idx = int(single_idx)
            class_names = [all_class_names[single_idx]]
            class_indices = [single_idx]
        except (ValueError, IndexError):
            class_names = all_class_names
            class_indices = list(range(num_classes))
    else:
        class_names   = all_class_names
        class_indices = list(range(num_classes))

    # Day labels: Mon–Sat for 6, Mon–Fri for 5, etc.
    _day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_labels = [_day_names[i] if i < len(_day_names) else f"Day {i+1}" for i in range(days)]

    # ── Style factories (new object per cell avoids openpyxl shared-style bugs) ─
    def hdr_fill():  return PatternFill("solid", fgColor="1F3864")
    def hdr_font():  return Font(color="FFFFFF", bold=True, size=11)
    def per_fill():  return PatternFill("solid", fgColor="D9E1F2")
    def per_font():  return Font(bold=True, size=10)
    def free_fill(): return PatternFill("solid", fgColor="FFF9C4")
    def lab_fill():  return PatternFill("solid", fgColor="E7F5FF")
    def norm_fill(): return PatternFill("solid", fgColor="FFFFFF")
    def mk_border(): return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"))
    def mk_center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for cls_idx, cls_name in zip(class_indices, class_names):
        ws = wb.create_sheet(title=f"Class {cls_name}"[:31])

        # Header row: Day | P1 | P2 | P3 | ...
        ws.row_dimensions[1].height = 26
        ws.column_dimensions["A"].width = 14

        for col, label in enumerate(["Day"] + [f"P{p+1}" for p in range(periods)]):
            c = ws.cell(row=1, column=col+1, value=label)
            c.fill      = hdr_fill()
            c.font      = hdr_font()
            c.alignment = mk_center()
            c.border    = mk_border()

        col_letters = list("BCDEFGHIJKLMNOPQRSTUVWXYZ")
        for p in range(periods):
            if p < len(col_letters):
                ws.column_dimensions[col_letters[p]].width = 24

        # Data rows — one row per day
        for d in range(days):
            row_num = d + 2
            ws.row_dimensions[row_num].height = 42

            # Day label cell
            dc = ws.cell(row=row_num, column=1, value=day_labels[d])
            dc.fill      = per_fill()
            dc.font      = per_font()
            dc.alignment = mk_center()
            dc.border    = mk_border()

            for p in range(periods):
                text, kind = _cell_text(timetable, cls_idx, d, p, periods)

                fill = {"free": free_fill(), "lab": lab_fill()}.get(kind, norm_fill())

                pc = ws.cell(row=row_num, column=p+2, value=text)
                pc.fill      = fill
                pc.alignment = mk_center()
                pc.border    = mk_border()
                pc.font      = Font(size=9, bold=(kind == "lab"),
                                    italic=(kind == "free"),
                                    color="6C757D" if kind == "free" else "000000")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"timetable_class_{class_names[0]}.xlsx" if len(class_names) == 1 else "timetable.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname
    )


# ─────────────────────────────────────────────────────────────────────────────
#  PDF DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/download/pdf")
def download_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    if not os.path.exists("generated_timetable.json") or \
       not os.path.exists("generated_metadata.json") or \
       not os.path.exists("temp_web_data.json"):
        return "No timetable found. Please generate one first.", 404

    with open("generated_timetable.json")  as f: timetable   = json.load(f)
    with open("generated_metadata.json")   as f: meta        = json.load(f)
    with open("temp_web_data.json")        as f: stored      = json.load(f)

    days        = meta["days"]
    periods     = meta["periods"]
    num_classes = meta["num_classes"]

    organized_keys = list(stored.get("organized", {}).keys())
    class_names = []
    for i in range(num_classes):
        class_names.append(organized_keys[i] if i < len(organized_keys) else str(i + 1))

    # Day labels: Mon–Sat for 6, Mon–Fri for 5, etc.
    _day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_labels = [_day_names[i] if i < len(_day_names) else f"Day {i+1}" for i in range(days)]

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm
    )

    styles  = getSampleStyleSheet()
    title_s = ParagraphStyle("ttl", parent=styles["Heading2"],
                             alignment=TA_CENTER, spaceAfter=6)
    cell_s  = ParagraphStyle("cel", parent=styles["Normal"],
                             fontSize=7, leading=9, alignment=TA_CENTER)
    free_s  = ParagraphStyle("fre", parent=styles["Normal"],
                             fontSize=7, leading=9, alignment=TA_CENTER,
                             textColor=colors.HexColor("#6C757D"))

    NAVY  = colors.HexColor("#1F3864")
    LBLUE = colors.HexColor("#D9E1F2")
    YFREE = colors.HexColor("#FFF9C4")
    BLAB  = colors.HexColor("#E7F5FF")
    WHITE = colors.white


    # Support single-class export via ?class_idx=N
    single_idx = request.args.get("class_idx", None)
    if single_idx is not None:
        try:
            single_idx = int(single_idx)
            class_indices = [single_idx]
            class_names   = [class_names[single_idx]] if single_idx < len(class_names) else class_names
        except (ValueError, IndexError):
            class_names   = [organized_keys[i] if i < len(organized_keys) else str(i+1) for i in range(num_classes)]
            class_indices = list(range(num_classes))
    else:
        class_indices = list(range(num_classes))

    story = []
    for cls_idx, cls_name in zip(class_indices, class_names):
        story.append(Paragraph(f"Class {cls_name} — Timetable", title_s))

        # Build table rows: header + one row per day
        header = ["Day"] + [f"P{p+1}" for p in range(periods)]
        rows   = [header]

        # Track which (row, col) cells need colour overrides
        free_cells = []
        lab_cells  = []

        for d in range(days):
            row = [Paragraph(day_labels[d], cell_s)]
            for p in range(periods):
                text, kind = _cell_text(timetable, cls_idx, d, p, periods)
                style = free_s if kind == "free" else cell_s
                row.append(Paragraph(text, style))
                if kind == "free":
                    free_cells.append((p+1, d+1))   # col, row
                elif kind == "lab":
                    lab_cells.append((p+1, d+1))
            rows.append(row)

        col_w = (27 * cm) / (periods + 1)
        t = Table(rows, colWidths=[col_w] * (periods + 1), repeatRows=1)

        ts = TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0),  NAVY),
            ("TEXTCOLOR",  (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",   (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0),  9),
            # Day column
            ("BACKGROUND", (0, 1), (0, -1),  LBLUE),
            ("FONTNAME",   (0, 1), (0, -1),  "Helvetica-Bold"),
            # All cells
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE",   (1, 1), (-1, -1), 8),
            ("ROWHEIGHT",  (0, 1), (-1, -1), 28),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ])

        # Apply per-cell colour overrides
        for (col_i, row_i) in free_cells:
            ts.add("BACKGROUND", (col_i, row_i), (col_i, row_i), YFREE)
        for (col_i, row_i) in lab_cells:
            ts.add("BACKGROUND", (col_i, row_i), (col_i, row_i), BLAB)

        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 0.8 * cm))

    doc.build(story)
    output.seek(0)
    fname = f"timetable_class_{class_names[0]}.pdf" if len(class_names) == 1 else "timetable.pdf"
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=fname
    )















# CLEANED: Only one version of success_summary using dynamic metadata
@app.route("/success-summary")
def success_summary():
    if not os.path.exists("generated_timetable.json") or not os.path.exists("generated_metadata.json"):
        return redirect(url_for('home'))

    with open("generated_timetable.json", "r") as f:
        timetable = json.load(f)
    with open("generated_metadata.json", "r") as f:
        meta = json.load(f)
    with open("final_schedule.json", "r") as f:
        final_data = json.load(f)
    with open("temp_web_data.json", "r") as f:
        stored = json.load(f)

    days        = meta['days']
    periods     = meta['periods']
    num_classes = meta['num_classes']

    class_names_raw = list(dict.fromkeys([row['class'] for row in final_data]))
    # Strip "Class " prefix stored in final_schedule
    class_names = [c.replace("Class ", "").strip() for c in class_names_raw]

    # ── Build teacher_slot_map: {teacher_name: ["classIdx-slotIdx", ...]} ──────
    # We need to know which teacher teaches each subject in each class
    organized = stored.get('organized', {})
    # subject->teacher lookup per class (also store normalized key for labs)
    subj_teacher = {}  # (class_idx, subject_lower_stripped) -> teacher_name
    for cidx, cname in enumerate(organized.keys()):
        for t in organized[cname]:
            # Primary key: exact subject name lowered
            key = (cidx, t['subject'].lower().strip())
            subj_teacher[key] = t['teacher']
            # Secondary key: strip "(lab N)" suffix so "Physics (Lab 1)" matches "Physics"
            import re as _re
            stripped = _re.sub(r'\s*\(lab[^)]*\)', '', t['subject'], flags=_re.IGNORECASE).lower().strip()
            if stripped != t['subject'].lower().strip():
                subj_teacher[(cidx, stripped)] = t['teacher']

    teacher_slot_map = {}   # teacher_name -> [classIdx-slotIdx]
    teacher_names_set = set()
    for cidx in range(num_classes):
        for day in range(days):
            for p in range(periods):
                si = day * periods + p
                try:
                    cell = timetable[si][cidx]
                except (IndexError, KeyError):
                    continue
                if not cell or cell == 0 or str(cell).lower() in ('free', 'f', '0'):
                    continue
                cell_str = str(cell).strip()
                import re as _re
                cell_norm = _re.sub(r'\s*\(lab[^)]*\)', '', cell_str, flags=_re.IGNORECASE).lower().strip()

                # Try exact match first, then stripped match, then prefix match
                key_exact   = (cidx, cell_str.lower().strip())
                key_stripped = (cidx, cell_norm)
                tname = subj_teacher.get(key_exact) or subj_teacher.get(key_stripped)
                if not tname:
                    # prefix fallback: find any subject that starts with first 6 chars
                    for (c2, subj), tn in subj_teacher.items():
                        if c2 == cidx and cell_norm.startswith(subj[:6]):
                            tname = tn
                            break
                if tname:
                    teacher_names_set.add(tname)
                    teacher_slot_map.setdefault(tname, []).append(f"{cidx}-{si}")

                # Also register sub-teachers for split blocks (e.g. "II Language" → aa, bb)
                # so they appear in the By Teacher view and are marked busy at these slots
                for bundle in (stored.get('auto_bundles', []) + stored.get('sync_groups', [])):
                    bname_lower = bundle.get('name', '').lower().strip()
                    if bname_lower in (cell_norm, cell_str.lower().strip()):
                        for m in bundle.get('members', []):
                            if int(m.get('classIdx', -1)) == cidx:
                                sub_t = m.get('teacherName', '')
                                if sub_t and sub_t != tname:
                                    teacher_names_set.add(sub_t)
                                    teacher_slot_map.setdefault(sub_t, []).append(f"{cidx}-{si}")

    teacher_names = sorted(teacher_names_set)

    # ── Build sync-group exempt set ───────────────────────────────────────────
    # Sync groups intentionally place the same teacher in multiple classes at
    # the same slot. Build (tname, slot_idx) pairs to skip in conflict detection.
    sync_exempt = set()   # {(teacher_name, slot_idx), ...}
    # Re-stamp auto_bundle classIdx using className before using them
    # (saved bundles can have stale indices; this ensures correct class matching)
    _ck = list(stored.get('organized', {}).keys())
    _fixed_abs = []
    for _ab in stored.get('auto_bundles', []):
        _fm = []
        for _m in _ab.get('members', []):
            _cn = _m.get('className', '').replace('Class ', '').strip()
            try: _idx = _ck.index(_cn)
            except ValueError: _idx = _m.get('classIdx', -1)
            _mf = dict(_m); _mf['classIdx'] = _idx; _fm.append(_mf)
        _fab = dict(_ab); _fab['members'] = _fm; _fixed_abs.append(_fab)
    # Include both UI-created sync_groups and auto-built bundles from Split rows
    sync_groups_stored = stored.get('sync_groups', []) + _fixed_abs
    for sg in sync_groups_stored:
        members = sg.get('members', [])
        if not members:
            continue
        # Exempt ALL member teachers from conflict detection — whether they're
        # shared across classes (multi-class sync) or are sub-teachers within
        # one class (intra-class split like II Language with eng/sans).
        # Without this, intra-class sub-teachers get double-counted as conflicts.
        all_member_teachers = {m.get('teacherName', '') for m in members if m.get('teacherName')}
        for tname_sg in all_member_teachers:
            if tname_sg in teacher_slot_map:
                for slot_str in teacher_slot_map[tname_sg]:
                    _, si_str = slot_str.split('-')
                    sync_exempt.add((tname_sg, int(si_str)))

    # ── Conflict checker: same teacher in 2 DIFFERENT classes at same slot ────
    # Use a SET of class indices so duplicate entries for the same class
    # (e.g. primary teacher + sub-teacher both added for class 0) don't falsely trigger.
    conflicts = []
    slot_teacher_classes = {}  # (tname, si) -> set of DISTINCT class_idxs
    for tname, slots in teacher_slot_map.items():
        for s in slots:
            cidx_str, si_str = s.split('-')
            key = (tname, int(si_str))
            slot_teacher_classes.setdefault(key, set()).add(int(cidx_str))
    for (tname, si), cidxs in slot_teacher_classes.items():
        if len(cidxs) > 1:  # only a real conflict if teacher in 2+ DIFFERENT classes
            # Skip if this (teacher, slot) is an intentional sync group assignment
            if (tname, si) in sync_exempt:
                continue
            for cidx in cidxs:
                conflicts.append([cidx, si])

    # teacher_map: {teacher_name: teacher_id} for JS
    teacher_map_js = {t['teacher']: t['teacher_id']
                      for cname in organized for t in organized[cname]}

    # ── Build sync_group_label_map for the template ─────────────────────────
    # Maps (cidx, subject_lower) -> bundle_name so the timetable can display
    # e.g. "2nd Language" instead of just "eng2" or "sans"
    sync_group_label_map = {}  # (cidx, subj_lower) -> bundle_display_name
    for sg in sync_groups_stored:  # already includes auto_bundles from above
        bname = sg.get('name', '')
        for m in sg.get('members', []):
            cidx_m = int(m.get('classIdx', -1))
            subj_m = (m.get('subject') or '').lower().strip()
            if cidx_m >= 0 and subj_m:
                sync_group_label_map[(cidx_m, subj_m)] = bname

    return render_template("success.html",
                           timetable=timetable,
                           num_classes=num_classes,
                           class_names=class_names,
                           num_days=days,
                           periods_per_day=periods,
                           teacher_names=teacher_names,
                           teacher_slot_map=teacher_slot_map,
                           teacher_map=teacher_map_js,
                           conflicts=conflicts,
                           sync_group_label_map={str(k): v for k, v in sync_group_label_map.items()})





# --- KEEP THIS VERSION (REPLACES THE TWO OLD ONES) ---
@app.route("/update-data", methods=["POST"])
def update_data():
    try:
        incoming_payload = request.get_json()
        web_data     = incoming_payload.get('table_data', [])
        config       = incoming_payload.get('config', {})
        split_groups = incoming_payload.get('split_groups', [])  # NEW: from Split rows

        # Create a mapping of teacher names to unique IDs
        all_teachers = sorted(list(set(row['teacher'] for row in web_data)))
        t_name_to_id = {name: i for i, name in enumerate(all_teachers)}

        # Build set of all (className, blockName) pairs that are split groups
        # so we can collapse sub-options into one block row per class
        split_block_seen = set()  # (className, blockName) already added as block row

        organized_classes = {}
        for row in web_data:
            c_name = row['class'].replace("Class ", "").strip()
            if c_name not in organized_classes:
                organized_classes[c_name] = []

            split_block = row.get('split_block', '').strip()

            if split_block:
                # This row is a sub-option of a split block.
                # For single-class splits: only add the BLOCK ITSELF once
                # (as a placeholder with the block name and correct hours).
                # Multiple sub-options in the same class all share those hours,
                # so we must not add them as separate subjects (would double/triple hours).
                # The sub-teacher info is carried in the auto_bundle for busy-marking.
                key = (c_name, split_block)
                if key not in split_block_seen:
                    split_block_seen.add(key)
                    # Use the first sub-option's teacher as the "primary" teacher
                    # for the block row — the bundle will mark all sub-teachers busy.
                    organized_classes[c_name].append({
                        "teacher":     row.get('teacher', 'Unknown'),
                        "teacher_id":  t_name_to_id.get(row.get('teacher'), 99),
                        "subject":     split_block,   # block name IS the subject in timetable
                        "hours":       int(row.get('periods', 0)),
                        "type":        "theory",
                        "continuous":  1,
                        "lab_no":      0,
                        "split_block": split_block,
                        "is_split_block": True,
                    })
                # Always skip adding the individual sub-option as its own subject row
                # — it would add extra hours to the class workload
                continue

            organized_classes[c_name].append({
                "teacher":    row.get('teacher', 'Unknown'),
                "teacher_id": t_name_to_id.get(row.get('teacher'), 99),
                "subject":    row.get('subject', 'General'),
                "hours":      int(row.get('periods', 0)),
                "type":       str(row.get('type', 'theory')).lower().strip(),
                "continuous": int(row.get('continuous', 1)),
                "lab_no":     int(row.get('lab_no', 0)),
                "split_block": '',
            })

        # ── Auto-build elective_bundles from split_groups ─────────────────────
        # split_groups format: [{blockName, className, hours, children:[{name,teacher}]}]
        # Group by blockName: same blockName across different classes → one bundle
        from collections import defaultdict
        block_classes = defaultdict(list)  # blockName -> [{ className, hours, children }]
        for sg in split_groups:
            block_classes[sg['blockName']].append(sg)

        auto_bundles = []
        for block_name, class_entries in block_classes.items():
            if len(class_entries) < 1:
                continue
            hours = class_entries[0]['hours']
            members = []
            # Get class index for each className
            class_keys = list(organized_classes.keys())
            for ce in class_entries:
                cname = ce['className'].replace('Class ', '').strip()
                try:
                    cidx = class_keys.index(cname)
                except ValueError:
                    continue
                for child in ce.get('children', []):
                    if not child.get('name') or not child.get('teacher'):
                        continue
                    members.append({
                        'classIdx':    cidx,
                        'className':   cname,
                        'subject':     child['name'],
                        'teacherName': child['teacher'],
                        'teacherId':   str(t_name_to_id.get(child['teacher'], 99)),
                        'hours':       hours
                    })
            # Create a bundle for ANY split group with 2+ sub-options,
            # even within a single class. The bundle forces all sub-subjects
            # (e.g. 'a' and 'b') to occupy the SAME slots, so the timetable
            # shows the block name ('II Language') not individual sub-subject names.
            # Multi-class bundles additionally sync across classes.
            if len(members) >= 2:
                unique_cidxs = {m['classIdx'] for m in members}
                # Rewrite each member's subject to the BLOCK NAME (e.g. 'II Language').
                # The subject_map only has the block name — sub-option names (eng/sans)
                # were collapsed into a single row and don't exist in subject_map.
                # Keep sub_subject for reference/display only.
                members_for_bundle = [
                    {**m, 'subject': block_name, 'sub_subject': m.get('subject', '')}
                    for m in members
                ]
                auto_bundles.append({
                    'name':          block_name,
                    'type':          'split',
                    'periodsPerWeek': hours,
                    'members':       members_for_bundle,
                })
                logging.info(f"Auto-bundle '{block_name}': {len(members_for_bundle)} members across {len(unique_cidxs)} class(es)")
            else:
                logging.info(f"Auto-bundle '{block_name}': fewer than 2 sub-options — skipping.")

        merge_groups = incoming_payload.get('merge_groups', [])

        session_data = {
            "organized":      organized_classes,
            "days":           int(config.get('days', 6)),
            "periods":        int(config.get('periods', 6)),
            "session_token":  str(__import__('uuid').uuid4()),
            "auto_bundles":   auto_bundles,
            "merge_groups":   merge_groups,
        }
        with open("temp_web_data.json", "w") as f:
            json.dump(session_data, f)

        return jsonify({"status": "success", "redirect": url_for('setup_fixed')})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500




# ── Load Verify (restore to verify/configure page) ────────────────────────────
@app.route("/load-verify", methods=["POST"])
def load_verify():
    """
    Stores a verify-page session (rows + days + periods + merge_groups)
    so /edit-schedule can render view_simple.html pre-populated.
    Called by:
      - upload page "Load Saved Data" (v2 saves with page='verify')
      - upload page "Enter Manually" (after class names, before subjects)
    """
    try:
        payload      = request.get_json()
        rows         = payload.get("rows", [])
        days         = int(payload.get("days", 6))
        periods      = int(payload.get("periods", 6))
        labs         = int(payload.get("labs", 2))
        merge_groups = payload.get("merge_groups", [])
        twd          = payload.get("temp_web_data")  # may be None for manual entry

        verify_session = {
            "rows":         rows,
            "days":         days,
            "periods":      periods,
            "labs":         labs,
            "merge_groups": merge_groups,
            "temp_web_data": twd,
        }
        with open("verify_session.json", "w") as f:
            json.dump(verify_session, f)

        return jsonify({"status": "success", "redirect": url_for("edit_schedule")})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Edit Schedule (verify page pre-populated from saved/manual data) ──────────
@app.route("/edit-schedule")
def edit_schedule():
    """
    Renders view_simple.html pre-populated from verify_session.json.
    This is used when the user loads a saved file or enters manually.
    The rows already include split_children so splits are shown correctly.
    """
    if not os.path.exists("verify_session.json"):
        return redirect(url_for("home"))

    with open("verify_session.json", "r") as f:
        vs = json.load(f)

    rows         = vs.get("rows", [])
    days         = int(vs.get("days", 6))
    periods      = int(vs.get("periods", 6))
    merge_groups = vs.get("merge_groups", [])

    import json as _json

    # Attach split_children_json to each row so the Jinja template can embed it
    for row in rows:
        children = row.get("split_children", [])
        row["split_children_json"] = _json.dumps(children)
        # Normalise type capitalisation
        t = str(row.get("type", "Theory"))
        row["type"] = t[0].upper() + t[1:].lower() if t else "Theory"
        # Remove "Class " prefix from class name if the user typed it already
        cn = row.get("class", "")
        if not cn.startswith("Class "):
            row["class"] = "Class " + cn

    return render_template(
        "view_simple.html",
        rows=rows,
        extracted_days=days,
        extracted_periods=periods,
        merge_groups_json=_json.dumps(merge_groups),
    )


# ── Load Save File ────────────────────────────────────────────────────────────
@app.route("/load-save", methods=["POST"])
def load_save():
    """
    Receives the temp_web_data blob from a downloaded save file,
    writes it to temp_web_data.json (same as /update-data does),
    and returns a new session_token so fixed_setup can match localStorage.
    """
    try:
        payload       = request.get_json()
        temp_web_data = payload.get("temp_web_data")
        if not temp_web_data:
            return jsonify({"status": "error", "message": "No temp_web_data in payload"}), 400

        # Issue a fresh session token — client will write this into localStorage
        # so fixed_setup.html trusts and loads the restored session data.
        import uuid
        new_token = str(uuid.uuid4())
        temp_web_data["session_token"] = new_token

        with open("temp_web_data.json", "w") as f:
            json.dump(temp_web_data, f)

        return jsonify({
            "status":        "success",
            "session_token": new_token,
            "redirect":      url_for("setup_fixed")
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# .................
@app.route("/setup-fixed")
def setup_fixed():
    if not os.path.exists("temp_web_data.json"):
        return redirect(url_for('home'))
        
    with open("temp_web_data.json", "r") as f:
        stored = json.load(f)
    
    # Defensive: always provide periods
    periods_value = stored.get('periods', 8)
    if not isinstance(periods_value, (int, float)):
        periods_value = 8

    # ── Re-stamp classIdx in auto_bundles using className ────────────────────
    # Saved bundles can have stale indices if class order differs from current
    # organized dict. Fix here so fixed_setup.html gets correct indices and
    # validateBeforeSolve doesn't block generation with false stale-index errors.
    class_keys_ordered = list(stored.get('organized', {}).keys())
    fixed_bundles = []
    for ab in stored.get('auto_bundles', []):
        fixed_members = []
        for m in ab.get('members', []):
            m_cn = m.get('className', '').replace('Class ', '').strip()
            try:
                fresh_idx = class_keys_ordered.index(m_cn)
            except ValueError:
                fresh_idx = m.get('classIdx', -1)  # keep as-is if not found
            fixed_m = dict(m)
            fixed_m['classIdx'] = fresh_idx
            fixed_members.append(fixed_m)
        fixed_ab = dict(ab)
        fixed_ab['members'] = fixed_members
        fixed_bundles.append(fixed_ab)
    # Also write back fixed bundles so run-final-solver gets clean data
    stored['auto_bundles'] = fixed_bundles
    with open("temp_web_data.json", "w") as f:
        json.dump(stored, f)

    return render_template(
        "fixed_setup.html",
        days=stored.get('days', 6),
        periods=periods_value,
        class_data=stored.get('organized', {}),
        session_token=stored.get('session_token', 'default'),
        auto_bundles=fixed_bundles,
        temp_web_data=stored
    )
@app.route("/run-final-solver", methods=["POST"])
def run_final_solver():
    try:
        payload        = request.get_json()
        fixed_data     = payload.get('fixed_slots', {})
        unavail_data   = payload.get('teacher_unavailability', {})
        elective_bundles = payload.get('elective_bundles', [])

        if not os.path.exists("temp_web_data.json"):
            return jsonify({"status": "error", "message": "Session expired. Please restart."}), 400

        with open("temp_web_data.json", "r") as f:
            stored = json.load(f)

        # Merge auto_bundles from split rows (persisted in session) with any
        # user-provided bundles from the sync group UI. User bundles take priority
        # (they override by name if the same blockName was also auto-built).
        auto_bundles = stored.get('auto_bundles', [])
        class_keys_for_bundles = list(stored.get('organized', {}).keys())
        if auto_bundles:
            existing_names = {b.get('name') for b in elective_bundles}
            for ab in auto_bundles:
                if ab.get('name') not in existing_names:
                    # Re-stamp classIdx from current class_keys using className
                    # (saved indices can be stale if class order changed)
                    fixed_members = []
                    for m in ab.get('members', []):
                        m_cn = m.get('className', '').replace('Class ', '').strip()
                        try:
                            fresh_idx = class_keys_for_bundles.index(m_cn)
                            m_fixed = dict(m)
                            m_fixed['classIdx'] = fresh_idx
                            fixed_members.append(m_fixed)
                        except ValueError:
                            fixed_members.append(m)  # keep as-is if not found
                    ab_fixed = dict(ab)
                    ab_fixed['members'] = fixed_members
                    elective_bundles.append(ab_fixed)

        # ── Convert merge_groups from verify page into extra elective_bundles ──
        # merge_groups = [{name, entries:[{className, blockName}]}]
        # Each merge group forces all its listed split blocks to share the same
        # K time slots, regardless of which class they belong to.
        merge_groups_stored = stored.get('merge_groups', [])
        class_keys = list(stored.get('organized', {}).keys())
        for mg in merge_groups_stored:
            mg_name = mg.get('name', 'Merge')
            entries = mg.get('entries', [])
            if len(entries) < 2:
                continue
            # Collect all sub-option members across all listed classes
            # Each entry: {className, blockName}  →  look up auto_bundle for that blockName
            mg_members = []
            mg_hours   = None
            existing_bundle_names = {b.get('name') for b in auto_bundles}
            for entry in entries:
                cn = entry.get('className', '').replace('Class ', '').strip()
                bn = entry.get('blockName', '').strip()
                # Find this class's index
                try:
                    cidx = class_keys.index(cn)
                except ValueError:
                    continue
                # Find the matching auto_bundle members for this class+blockName
                # Use className (reliable) not classIdx (can be stale from old saves)
                for ab in auto_bundles:
                    if ab.get('name') == bn:
                        for m in ab.get('members', []):
                            m_cn = m.get('className', '').replace('Class ', '').strip()
                            if m_cn == cn:
                                # Re-stamp classIdx from current class_keys so solver sees correct index
                                m_fixed = dict(m)
                                m_fixed['classIdx'] = cidx
                                mg_members.append(m_fixed)
                                if mg_hours is None:
                                    mg_hours = ab.get('periodsPerWeek', 3)
            if len({m.get('classIdx') for m in mg_members}) >= 2 and mg_members:
                # Only add if not already an auto_bundle with the same name
                if mg_name not in {b.get('name') for b in elective_bundles}:
                    elective_bundles.append({
                        'name':           mg_name,
                        'type':           'merged',
                        'periodsPerWeek': mg_hours or 3,
                        'members':        mg_members,
                    })
                    logging.info(f"Merge group '{mg_name}': {len(mg_members)} members across "
                                 f"{len({m['classIdx'] for m in mg_members})} classes")

        from adapter import build_final_inputs 
        
        (No_of_classes, t_list, c_theory, l_periods, subj_map) = build_final_inputs(
            {"classes": stored['organized']}, 
            stored['days'], 
            stored['periods'], 
            fixed_data
        )

        # ── Convert sync_groups / elective_bundles from frontend to solver format ─
        solver_bundles = []
        if elective_bundles:
            for b in elective_bundles:
                solver_bundles.append({
                    "name":             b.get("name", ""),
                    "type":             b.get("type", "split"),
                    "periodsPerWeek":   int(b.get("periodsPerWeek", 1)),
                    "members":          b.get("members", []),
                    # Legacy fields kept for backward compat with old backtracker path
                    "assignments":      b.get("assignments", {}),
                })

        # --- DEBUG LOGGING ---
        debug_payload = {
            "No_of_classes": No_of_classes,
            "days": stored['days'],
            "periods": stored['periods'],
            "teacher_list": t_list,
            "class_theory_workload": c_theory,
            "lab_periods": l_periods,
            "subject_map": {str(k): v for k, v in subj_map.items()},
            "fixed_periods": fixed_data,
            "elective_bundles": solver_bundles
        }
        with open("solver_input_debug.json", "w") as f:
            json.dump(debug_payload, f, indent=4)

        final_timetable = generate_timetable_with_retry(
            No_of_classes, stored['days'], stored['periods'], t_list,
            c_theory, l_periods, subj_map,
            fixed_periods=fixed_data,
            teacher_unavailability=unavail_data,
            elective_bundles=solver_bundles
        )

        if final_timetable:
            # 1. Save metadata for the success page
            with open("generated_metadata.json", "w") as f:
                json.dump({
                    "days": stored['days'], 
                    "periods": stored['periods'], 
                    "num_classes": No_of_classes
                }, f)
            
            # 2. Save the actual timetable
            with open("generated_timetable.json", "w") as f:
                json.dump(final_timetable, f)

            # 3. Create the 'final_schedule.json' that success_summary expects
            flat_rows = []
            for c_name, teachers in stored['organized'].items():
                for t in teachers:
                    flat_rows.append({"class": f"Class {c_name}", "teacher": t['teacher']})
            with open("final_schedule.json", "w") as f:
                json.dump(flat_rows, f)

            # 4. Persist sync groups into temp_web_data so success_summary can
            #    exempt intentional shared-teacher slots from conflict detection
            stored['sync_groups'] = solver_bundles
            with open("temp_web_data.json", "w") as f:
                json.dump(stored, f)

            return jsonify({"status": "success", "redirect": url_for('success_summary')})
        
        # ── Smart solver failure diagnostics ──────────────────────────────────
        report_lines = []
        days        = stored['days']
        periods_day = stored['periods']
        total_slots = days * periods_day
        organized   = stored['organized']

        # 1. Per-class overload
        for cname, teachers in organized.items():
            theory_hrs = sum(int(t.get('hours', 0)) for t in teachers if t.get('type','theory').lower() != 'lab')
            lab_hrs    = sum(int(t.get('hours', 0)) for t in teachers if t.get('type','').lower() == 'lab')
            total_hrs  = theory_hrs + lab_hrs
            if total_hrs > total_slots:
                over = total_hrs - total_slots
                report_lines.append(
                    f"📚 <b>Class {cname}</b> has <b>{total_hrs} hours</b> but only "
                    f"<b>{total_slots} slots</b> available ({over} hour(s) too many). "
                    f"Remove or reduce a subject."
                )

        # 2. Teacher overload — total hours across all classes vs available slots
        teacher_hours = {}
        for cname, teachers in organized.items():
            for t in teachers:
                tname = t.get('teacher', '')
                hours = int(t.get('hours', 0))
                teacher_hours.setdefault(tname, 0)
                teacher_hours[tname] += hours
        for tname, total in teacher_hours.items():
            if total > total_slots:
                over = total - total_slots
                report_lines.append(
                    f"👤 <b>{tname}</b> is assigned <b>{total} hours total</b> across all classes "
                    f"but only {total_slots} slots exist per week ({over} too many). "
                    f"Reduce this teacher's hours or split across different teachers."
                )

        # 3. Fixed slot overcommitment — count fixed slots per class
        fixed_counts = {}
        for cls_str, slots in fixed_data.items():
            count = sum(1 for s in slots.values()
                        if s.get('teacher_id', '__none__') != '__none__' and s.get('label'))
            if count:
                fixed_counts[cls_str] = count
        for cls_str, count in fixed_counts.items():
            cidx = int(cls_str)
            cname = list(organized.keys())[cidx] if cidx < len(organized) else cls_str
            avail = total_slots - count
            theory_needed = sum(int(t.get('hours',0)) for t in organized.get(cname,[])
                                if t.get('type','theory').lower() != 'lab')
            if theory_needed > avail:
                report_lines.append(
                    f"📌 <b>Class {cname}</b>: {count} fixed slots leave only {avail} free slots "
                    f"but theory subjects need {theory_needed}. Remove some fixed slots."
                )

        # 4. Sync group problems
        for b in solver_bundles:
            bname   = b.get('name', 'unnamed')
            k       = int(b.get('periodsPerWeek', 1))
            members = b.get('members', [])
            if len(members) < 2:
                report_lines.append(
                    f"🔗 Sync group <b>\"{bname}\"</b> has fewer than 2 members — skipped by solver."
                )
            # Check every member: does classIdx actually contain that subject?
            class_keys = list(organized.keys())
            for m in members:
                cidx      = int(m.get('classIdx', -1))
                subj_name = (m.get('subject') or '').strip()
                tname_m   = m.get('teacherName', '')
                if cidx < 0 or cidx >= len(class_keys):
                    report_lines.append(
                        f"🔗 Sync group <b>\"{bname}\"</b>: classIdx <b>{cidx}</b> is out of range "                        f"(only {len(class_keys)} classes exist: indices 0–{len(class_keys)-1}). "                        f"Re-create the sync group — pick subjects from the correct class rows in the dropdown."
                    )
                    continue
                cname_m   = class_keys[cidx]
                stored_cname = (m.get('className') or '').strip()
                class_subjs = [t.get('subject','') for t in organized.get(cname_m, [])]
                # Also include split-block sub-subjects in the valid subject list
                class_subjs_all = class_subjs  # same list, split_block subjects are in organized too
                if subj_name not in class_subjs:
                    # Detect stale classIdx: the name stored in the member doesn't match
                    # what's at that index now — this is a stale-localStorage problem.
                    if stored_cname and stored_cname != cname_m:
                        # Try to find the right index for the stored class name
                        correct_idx = class_keys.index(stored_cname) if stored_cname in class_keys else -1
                        correct_subjs = [t.get('subject','') for t in organized.get(stored_cname, [])]
                        if correct_idx >= 0 and subj_name in correct_subjs:
                            report_lines.append(
                                f"🔗 Sync group <b>\"{bname}\"</b>: member says class <b>\"{stored_cname}\"</b> "
                                f"but classIdx <b>{cidx}</b> points to <b>\"{cname_m}\"</b> instead. "
                                f"This is stale data from a previous session. "
                                f"<b>Fix:</b> On the Class-Specific Setup page, open the Sync Groups panel, "
                                f"delete group <b>\"{bname}\"</b>, then re-create it — it should be "
                                f"auto-populated from your Split rows. Or click the page back and forward to reload."
                            )
                        else:
                            report_lines.append(
                                f"🔗 Sync group <b>\"{bname}\"</b>: subject <b>\"{subj_name}\"</b> "
                                f"does not exist in <b>Class {cname_m}</b> (index {cidx}). "
                                f"That class has: {', '.join(class_subjs[:6])}. "
                                f"Delete this sync group and re-add it from the correct class rows."
                            )
                    else:
                        report_lines.append(
                            f"🔗 Sync group <b>\"{bname}\"</b>: subject <b>\"{subj_name}\"</b> "
                            f"does not exist in <b>Class {cname_m}</b> (index {cidx}). "
                            f"That class has: {', '.join(class_subjs[:6])}. "
                            f"Delete this sync group and re-add using the correct class rows in the dropdown."
                        )
            # Check if any member's teacher is overloaded with sync slots
            teacher_sync_load = {}
            for m in members:
                tid = str(m.get('teacherId',''))
                tname = m.get('teacherName','?')
                teacher_sync_load.setdefault(tname, 0)
                teacher_sync_load[tname] += k
            for tname, sync_hrs in teacher_sync_load.items():
                total_for_teacher = teacher_hours.get(tname, 0)
                if sync_hrs > total_slots:
                    report_lines.append(
                        f"🔗 Sync group <b>\"{bname}\"</b>: teacher <b>{tname}</b> would need "
                        f"{sync_hrs} slots just for sync assignments but only {total_slots} slots exist."
                    )

        # 5. Teacher unavailability too restrictive
        if unavail_data:
            for tid_str, blocked_slots in unavail_data.items():
                # Find teacher name
                tname = tid_str
                for cname, teachers in organized.items():
                    for t in teachers:
                        if str(t.get('teacher_id','')) == tid_str:
                            tname = t.get('teacher', tid_str)
                            break
                total_blocked = len(blocked_slots)
                avail_slots   = total_slots - total_blocked
                needed = teacher_hours.get(tname, 0)
                if needed > avail_slots:
                    report_lines.append(
                        f"🚫 <b>{tname}</b> has {total_blocked} unavailable slots, leaving "
                        f"{avail_slots} free — but needs {needed} teaching slots. "
                        f"Reduce unavailability or reduce their hours."
                    )

        # Phantom class detection: class with >80% free slots is likely a parsing artifact
        total_s = days * periods_day
        for cname, teachers in organized.items():
            total_hours = sum(int(t.get('hours', 0)) for t in teachers if t.get('type','theory').lower() != 'lab')
            lab_hours = sum(int(t.get('hours', 0)) for t in teachers if t.get('type','').lower() == 'lab')
            real_hours = total_hours + lab_hours
            if real_hours < total_s * 0.2:  # Less than 20% real subjects
                free_hrs = total_s - real_hours
                report_lines.append(
                    f"🔍 <b>Class {cname}</b> has only <b>{real_hours} real subject hours</b> "                    f"({free_hrs} free slots out of {total_s}). "                    f"This is likely a <b>PDF extraction artifact</b> — check if this is a real class "                    f"or leftover data from the last page of the PDF. "                    f"If it's not a real class, delete all its rows in Data Verification."
                )

        # OR-Tools missing warning
        try:
            from ortools.sat.python import cp_model as _cp
        except ImportError:
            report_lines.append(
                f"⚡ <b>OR-Tools is not installed.</b> The backtracking solver is much slower "                f"and may fail on inputs this size. "                f"Run <code>pip install ortools</code> and restart the app to use the fast CP-SAT solver."
            )

        if not report_lines:
            report_lines.append(
                "🤔 No obvious overload found. Possible causes:<br>"
                "• Fixed slots are blocking too many combinations for the solver to fit everything.<br>"
                "• Sync group constraints conflict with teacher availability.<br>"
                "• A teacher teaches many classes and their slots are tightly constrained.<br>"
                "<b>Try:</b> removing some fixed slots, relaxing unavailability, or reducing sync group size."
            )

        conflict_report = "<br><br>".join(report_lines)
        return jsonify({"status": "error", "message": "Solver could not find a valid timetable.", "conflict_report": conflict_report})
    except Exception as e:
        import traceback
        print(traceback.format_exc()) 
        return jsonify({"status": "error", "message": str(e)}), 500

        


@app.route("/swap-slots", methods=["POST"])
def swap_slots():
    try:
        data      = request.get_json()
        class_idx = int(data['class_idx'])
        si1       = int(data['slot1'])
        si2       = int(data['slot2'])

        if not os.path.exists("generated_timetable.json"):
            return jsonify({"status": "error", "message": "No timetable found"}), 404

        with open("generated_timetable.json") as f:
            timetable = json.load(f)

        # Swap the two slots for the given class
        timetable[si1][class_idx], timetable[si2][class_idx] = \
            timetable[si2][class_idx], timetable[si1][class_idx]

        with open("generated_timetable.json", "w") as f:
            json.dump(timetable, f)

        return jsonify({"status": "success"})
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)